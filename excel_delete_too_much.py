import openpyxl
import re

filename = '4.xlsx'
wb = openpyxl.load_workbook(filename)
sheet = wb['Лист1']
m = 0
f = sheet.max_row
k = 0
for n in range(sheet.max_row):
    k =+ 1
    c = sheet[f'B{n + 1}'].value
    d = sheet[f'A{n + 1}'].value
    a = str(c)
    b = re.sub(r' ', '', a)
    b = re.sub(r',$', '', b)
    b = re.split(r',', b)
    if len(b) > 1:
        print(b)
        print(n+1)
        sheet.delete_rows(n + 1)
        n = n - 1
        m += 1

wb.save("4.xlsx")
print(m)
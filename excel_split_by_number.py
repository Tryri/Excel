import openpyxl
import re

filename = '1.xlsx'
wb = openpyxl.load_workbook(filename)
sheet = wb.active
f = sheet.max_row
for n in range(sheet.max_row):
    # if n.value != None: # range(sheet.max_row)
    a = sheet[f'B{n + 1}'].value  # Номера домов
    b = sheet[f'A{n + 1}'].value  # Улицы
    a = str(a)
    a = re.sub(r',$', '', a)
    a = re.split(r',', a)
    # print(n.row)
    print(a)
    if len(a) > 1:
        for k in a:
            print(b, k)
            sheet.append([b, k])
        # sheet.delete_rows(n.row)



wb.save("1.xlsx")